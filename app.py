import os
import gc
import re
import json
import time
import base64
import random
import hashlib
import bcrypt
import openpyxl
import pandas as pd
import streamlit as st
from io import BytesIO
from PIL import Image
from google import genai
from google.genai import types
from rapidfuzz import process, utils
from tenacity import retry, stop_after_attempt, wait_exponential, retry_if_exception
from concurrent.futures import ThreadPoolExecutor
from sqlalchemy import create_engine, text

# Optional Imports with Fallbacks
try:
    import fitz  # PyMuPDF
    PYMUPDF_AVAILABLE = True
except ImportError:
    PYMUPDF_AVAILABLE = False

try:
    from groq import Groq
    GROQ_AVAILABLE = True
except ImportError:
    GROQ_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# --- PAGE CONFIGURATION ---
st.set_page_config(
    page_title="Universal OS | Multi-Store AI SaaS",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded"
)

# --- GLOBAL SESSION STATE ---
if "authenticated" not in st.session_state:
    st.session_state["authenticated"] = False
if "user_store" not in st.session_state:
    st.session_state["user_store"] = None
if "ocr_file_hash_cache" not in st.session_state:
    st.session_state["ocr_file_hash_cache"] = {}
if "catalog_df" not in st.session_state:
    st.session_state["catalog_df"] = None

# --- DESIGN SYSTEM ---
st.markdown("""
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap" rel="stylesheet">
<style>
    :root {
        --uos-primary: #4F46E5;
        --uos-primary-600: #4338CA;
        --uos-primary-700: #3730A3;
        --uos-primary-50: #EEF2FF;
        --uos-accent: #10B981;
        --uos-canvas: #F8FAFC;
        --uos-surface: #FFFFFF;
        --uos-surface-2: #F1F5F9;
        --uos-border: #E2E8F0;
        --uos-text: #0F172A;
        --uos-text-secondary: #475569;
        --uos-text-muted: #64748B;
        --uos-radius: 12px;
    }
    html, body, [class*="css"], .stApp, [data-testid="stAppViewContainer"] {
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif !important;
        background: var(--uos-canvas) !important;
        color: var(--uos-text) !important;
    }
    [data-testid="stHeader"] { background: transparent !important; }
    .block-container { padding-top: 1.5rem !important; padding-bottom: 3rem !important; max-width: 1400px; }
    div[data-testid="InputInstructions"] { display: none !important; }
    h1, h2, h3, h4 { font-family: 'Inter', sans-serif !important; color: var(--uos-text) !important; font-weight: 700 !important; }
    [data-testid="stSidebar"] { background: var(--uos-surface) !important; border-right: 1px solid var(--uos-border); }
    .uos-store-avatar { display: flex; align-items: center; gap: 12px; padding: 4px 2px 14px 2px; }
    .uos-store-avatar .uos-avatar-circle {
        width: 42px; height: 42px; border-radius: 12px;
        background: linear-gradient(135deg, var(--uos-primary) 0%, var(--uos-primary-700) 100%);
        color: #FFFFFF; display: flex; align-items: center; justify-content: center; font-weight: 700;
    }
    .uos-store-pill {
        display: inline-flex; align-items: center; gap: 8px; padding: 6px 14px 6px 10px;
        background: var(--uos-primary-50); color: var(--uos-primary); border: 1px solid #E0E7FF;
        border-radius: 999px; font-size: 0.82rem; font-weight: 600;
    }
    .uos-store-pill .uos-dot { width: 8px; height: 8px; border-radius: 50%; background: var(--uos-accent); }
    .stTabs [data-baseweb="tab-list"] { gap: 4px; border-bottom: 1px solid var(--uos-border); background: transparent !important; }
    .stTabs [data-baseweb="tab"] { padding: 12px 18px !important; font-weight: 500 !important; color: var(--uos-text-muted) !important; }
    .stTabs [data-baseweb="tab"][aria-selected="true"] { color: var(--uos-primary) !important; font-weight: 600 !important; border-bottom-color: var(--uos-primary) !important; }
    [data-testid="stMetric"] { background: var(--uos-surface) !important; border: 1px solid var(--uos-border) !important; border-radius: var(--uos-radius) !important; padding: 18px !important; }
    .stButton > button, .stDownloadButton > button, .stFormSubmitButton > button {
        border-radius: var(--uos-radius) !important; font-weight: 600 !important; border: 1px solid var(--uos-border) !important;
    }
    .stButton > button[kind="primary"] { background: var(--uos-primary) !important; color: #FFFFFF !important; }
</style>
""", unsafe_allow_html=True)

# --- DATABASE ENGINE SETUP ---
@st.cache_resource
def get_db_engine():
    db_url = st.secrets["SUPABASE_DB_URL"]
    return create_engine(db_url, pool_pre_ping=True, pool_size=10, max_overflow=20)

def sanitize_store_slug(name):
    return "".join([c if c.isalnum() else "_" for c in name.strip()]).lower()

@st.cache_data(ttl=3600)
def get_or_create_store_id(store_slug: str, display_name: str = None) -> int:
    engine = get_db_engine()
    slug = sanitize_store_slug(store_slug)
    if not display_name:
        display_name = store_slug.replace("_", " ").title()
    
    with engine.begin() as conn:
        result = conn.execute(
            text("SELECT id FROM stores WHERE slug = :slug"),
            {"slug": slug}
        ).fetchone()
        
        if result:
            return result[0]
        
        insert_res = conn.execute(
            text("INSERT INTO stores (slug, display_name) VALUES (:slug, :display_name) RETURNING id"),
            {"slug": slug, "display_name": display_name}
        )
        return insert_res.fetchone()[0]

@st.cache_data(ttl=600)
def get_store_phone(store_slug: str) -> str:
    engine = get_db_engine()
    try:
        with engine.connect() as conn:
            res = conn.execute(
                text("SELECT phone FROM stores WHERE slug = :slug"),
                {"slug": store_slug}
            ).fetchone()
            if res and res[0]:
                return res[0]
    except Exception:
        pass
    return ""

def save_store_phone(store_slug: str, phone: str):
    engine = get_db_engine()
    try:
        with engine.begin() as conn:
            conn.execute(
                text("UPDATE stores SET phone = :phone WHERE slug = :slug"),
                {"phone": phone.strip(), "slug": store_slug}
            )
        get_store_phone.clear()
    except Exception:
        pass

# --- PERMANENT DATABASE DUPLICATE CHECKER ---
def is_duplicate_invoice_db(store_id: int, vendor_name: str, invoice_number: str) -> bool:
    if not invoice_number or not vendor_name:
        return False
    engine = get_db_engine()
    v_clean = vendor_name.strip().upper()
    inv_clean = invoice_number.strip().upper()
    
    query = text("""
        SELECT id FROM processed_invoices 
        WHERE store_id = :store_id 
          AND UPPER(vendor_name) = :vendor 
          AND UPPER(invoice_number) = :inv
    """)
    try:
        with engine.connect() as conn:
            res = conn.execute(query, {"store_id": store_id, "vendor": v_clean, "inv": inv_clean}).fetchone()
            return res is not None
    except Exception:
        return False

def log_invoice_to_db(store_id: int, vendor_name: str, invoice_number: str, invoice_date: str, grand_total: float):
    if not invoice_number or not vendor_name:
        return
    engine = get_db_engine()
    v_clean = vendor_name.strip().upper()
    inv_clean = invoice_number.strip().upper()
    
    query = text("""
        INSERT INTO processed_invoices (store_id, vendor_name, invoice_number, invoice_date, grand_total)
        VALUES (:store_id, :vendor, :inv, :inv_date, :total)
        ON CONFLICT (store_id, vendor_name, invoice_number) DO NOTHING
    """)
    try:
        with engine.begin() as conn:
            conn.execute(query, {
                "store_id": store_id,
                "vendor": v_clean,
                "inv": inv_clean,
                "inv_date": invoice_date or "",
                "total": grand_total
            })
    except Exception:
        pass

# --- AUTHENTICATION ---
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def check_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def register_user(store_name: str, email: str, password: str):
    engine = get_db_engine()
    slug = sanitize_store_slug(store_name)
    hashed_pwd = hash_password(password)
    
    with engine.begin() as conn:
        existing = conn.execute(
            text("SELECT id FROM stores WHERE email = :email OR slug = :slug"),
            {"email": email.strip().lower(), "slug": slug}
        ).fetchone()
        
        if existing:
            return False, "Store name or Email already registered!"
            
        conn.execute(
            text("""
                INSERT INTO stores (slug, display_name, email, password)
                VALUES (:slug, :display_name, :email, :password)
            """),
            {
                "slug": slug,
                "display_name": store_name.strip(),
                "email": email.strip().lower(),
                "password": hashed_pwd
            }
        )
    return True, "Store registered successfully! Please log in."

def authenticate_user(email: str, password: str):
    engine = get_db_engine()
    with engine.connect() as conn:
        user = conn.execute(
            text("SELECT slug, display_name, password, phone FROM stores WHERE email = :email"),
            {"email": email.strip().lower()}
        ).fetchone()
        
        if user and user[2] and check_password(password, user[2]):
            return {
                "slug": user[0], 
                "display_name": user[1],
                "phone": user[3] if len(user) > 3 and user[3] else ""
            }
    return None

def get_user_by_slug(slug: str):
    engine = get_db_engine()
    try:
        with engine.connect() as conn:
            user = conn.execute(
                text("SELECT slug, display_name, phone FROM stores WHERE slug = :slug"),
                {"slug": slug.strip().lower()}
            ).fetchone()
            
            if user:
                return {
                    "slug": user[0],
                    "display_name": user[1],
                    "phone": user[2] if len(user) > 2 and user[2] else ""
                }
    except Exception:
        pass
    return None

def reset_user_password(email: str, new_password: str):
    engine = get_db_engine()
    hashed_pwd = hash_password(new_password)
    
    with engine.begin() as conn:
        user = conn.execute(
            text("SELECT id FROM stores WHERE email = :email"),
            {"email": email.strip().lower()}
        ).fetchone()
        
        if not user:
            return False, "No store registered with this business email address."
            
        conn.execute(
            text("UPDATE stores SET password = :password WHERE email = :email"),
            {"password": hashed_pwd, "email": email.strip().lower()}
        )
    return True, "Password updated successfully!"

# --- FAST SESSION RESOLUTION ---
if not st.session_state["authenticated"]:
    url_session = st.query_params.get("session", None)
    if url_session:
        if isinstance(url_session, list):
            url_session = url_session[0]
        
        store_data = get_user_by_slug(str(url_session))
        if store_data:
            st.session_state["authenticated"] = True
            st.session_state["user_store"] = store_data

# --- LOGIN SCREEN ---
if not st.session_state["authenticated"]:
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown("""
            <div class="uos-auth-hero">
                <div class="uos-auth-mark">⚡</div>
                <h1 class="uos-auth-title">Universal OS</h1>
                <p class="uos-auth-tagline">Commercial AI Invoice Intake &amp; ERP Synchronizer</p>
            </div>
        """, unsafe_allow_html=True)
        st.write("")
        
        auth_tab1, auth_tab2, auth_tab3 = st.tabs(["🔒 Partner Login", "✨ Register New Store", "🔑 Forgot Password"])
        
        with auth_tab1:
            with st.form("login_form"):
                login_email = st.text_input("Store Email", key="login_email_input")
                login_password = st.text_input("Password", type="password", key="login_pass_input")
                submit_login = st.form_submit_button("Log In", use_container_width=True, type="primary")
                
                if submit_login:
                    if login_email and login_password:
                        store_data = authenticate_user(login_email, login_password)
                        if store_data:
                            st.session_state["authenticated"] = True
                            st.session_state["user_store"] = store_data
                            st.query_params["session"] = store_data["slug"]
                            st.rerun()
                        else:
                            st.error("Invalid email or password.")
                    else:
                        st.warning("Please fill in both email and password.")

        with auth_tab2:
            with st.form("register_form"):
                reg_store = st.text_input("Store Name")
                reg_email = st.text_input("Business Email")
                reg_password = st.text_input("Create Password", type="password")
                submit_reg = st.form_submit_button("Create Account & Store Environment", use_container_width=True)
                
                if submit_reg:
                    if reg_store and reg_email and reg_password:
                        success, msg = register_user(reg_store, reg_email, reg_password)
                        if success: st.success(msg)
                        else: st.error(msg)
                    else: st.warning("All fields are required.")

        with auth_tab3:
            with st.form("reset_password_form"):
                reset_email = st.text_input("Registered Business Email")
                new_pwd = st.text_input("New Password", type="password")
                confirm_pwd = st.text_input("Confirm New Password", type="password")
                submit_reset = st.form_submit_button("Reset Password", use_container_width=True)
                
                if submit_reset:
                    if not reset_email or not new_pwd: st.warning("Please enter your email and a new password.")
                    elif new_pwd != confirm_pwd: st.error("Passwords do not match!")
                    else:
                        success, msg = reset_user_password(reset_email, new_pwd)
                        if success: st.success(msg)
                        else: st.error(msg)
    st.stop()

# Active user details
selected_store_slug = st.session_state["user_store"]["slug"]
ACTIVE_STORE_DISPLAY = st.session_state["user_store"]["display_name"].upper()
ACTIVE_STORE_ID = get_or_create_store_id(selected_store_slug)

# --- SIDEBAR ---
_display_name = st.session_state['user_store']['display_name']
_initials = "".join([w[0].upper() for w in _display_name.split()[:2] if w]) or "US"
st.sidebar.markdown(f"""
    <div class="uos-store-avatar">
        <div class="uos-avatar-circle">{_initials}</div>
        <div class="uos-avatar-meta">
            <div class="uos-avatar-name">{_display_name}</div>
            <div class="uos-avatar-slug">{selected_store_slug}</div>
        </div>
    </div>
""", unsafe_allow_html=True)
st.sidebar.divider()

if st.sidebar.button("🚪 Logout", use_container_width=True):
    st.session_state["authenticated"] = False
    st.session_state["user_store"] = None
    st.query_params.clear()
    if "parsed_df" in st.session_state:
        del st.session_state["parsed_df"]
    st.rerun()

st.sidebar.divider()

# --- MULTI-KEY & MULTI-MODEL FAIL-SAFE API POOL SETUP ---
def get_api_key_pool():
    keys = []
    for i in range(1, 6):
        key = st.secrets.get(f"GEMINI_API_KEY_{i}") or os.environ.get(f"GEMINI_API_KEY_{i}")
        if key: keys.append(key)
    if not keys:
        single_key = st.secrets.get("GEMINI_API_KEY") or os.environ.get("GEMINI_API_KEY")
        if single_key: keys.append(single_key)
    return keys

API_KEYS_POOL = get_api_key_pool()
if not API_KEYS_POOL:
    st.error("⚠️ Gemini API Key missing.")
    st.stop()

# --- HEADER SECTION ---
header_left, header_right = st.columns([3, 1.5])
with header_left:
    st.markdown("""
        <div style="display:flex;align-items:center;gap:14px;">
            <div style="width:44px;height:44px;border-radius:14px;
                        background:linear-gradient(135deg,#4F46E5 0%,#3730A3 100%);
                        color:#fff;display:flex;align-items:center;justify-content:center;
                        font-size:1.4rem;box-shadow:0 8px 20px rgba(79,70,229,0.25);flex-shrink:0;">⚡</div>
            <div>
                <h1 style="margin:0;font-size:1.75rem;font-weight:700;letter-spacing:-0.03em;line-height:1.2;">Universal OS</h1>
                <div style="color:var(--uos-text-muted);font-size:0.85rem;margin-top:2px;">Commercial Multi-Store AI Purchase Intake &amp; Inventory Synchronizer</div>
            </div>
        </div>
    """, unsafe_allow_html=True)

with header_right:
    st.markdown(f"""
        <div style="display:flex;justify-content:flex-end;align-items:center;height:100%;padding-top:12px;">
            <div class="uos-store-pill">
                <span class="uos-dot"></span>
                <span class="uos-pill-label">Active Store</span>
                <span>{ACTIVE_STORE_DISPLAY}</span>
            </div>
        </div>
    """, unsafe_allow_html=True)

st.divider()

# --- STORE LOADERS ---
@st.cache_data(ttl=3600)
def load_json_memory(store_slug: str) -> dict:
    engine = get_db_engine()
    store_id = get_or_create_store_id(store_slug)
    query = "SELECT raw_name, mapped_sku FROM vendor_mappings WHERE store_id = :store_id"
    try:
        with engine.connect() as conn:
            df = pd.read_sql(text(query), conn, params={"store_id": store_id})
        if df.empty: return {}
        return dict(zip(df["raw_name"], df["mapped_sku"]))
    except Exception:
        return {}

def save_json_memory(store_slug: str, memory_dict: dict):
    if not memory_dict: return
    engine = get_db_engine()
    store_id = get_or_create_store_id(store_slug)
    records = [{"store_id": store_id, "raw_name": k, "mapped_sku": v} for k, v in memory_dict.items()]
    with engine.begin() as conn:
        conn.execute(
            text("""
                INSERT INTO vendor_mappings (store_id, raw_name, mapped_sku)
                VALUES (:store_id, :raw_name, :mapped_sku)
                ON CONFLICT (store_id, raw_name) 
                DO UPDATE SET mapped_sku = EXCLUDED.mapped_sku
            """),
            records
        )
    load_json_memory.clear()

@st.cache_data(ttl=3600)
def load_master(store_slug: str) -> pd.DataFrame:
    engine = get_db_engine()
    store_id = get_or_create_store_id(store_slug)
    query = """
        SELECT official_sku_name as "Official_SKU_Name", 
               category as "Category", 
               default_unit as "Default_Unit", 
               gst_rate as "GST_Rate", 
               selling_price as "Selling_Price"
        FROM master_skus
        WHERE store_id = :store_id
    """
    try:
        with engine.connect() as conn:
            df = pd.read_sql(text(query), conn, params={"store_id": store_id})
        if df.empty: return pd.DataFrame(columns=["Official_SKU_Name", "Category", "Default_Unit", "GST_Rate", "Selling_Price"])
        return df
    except Exception:
        return pd.DataFrame(columns=["Official_SKU_Name", "Category", "Default_Unit", "GST_Rate", "Selling_Price"])

def save_master(df: pd.DataFrame, store_slug: str):
    engine = get_db_engine()
    store_id = get_or_create_store_id(store_slug)
    with engine.begin() as conn:
        conn.execute(text("DELETE FROM master_skus WHERE store_id = :store_id"), {"store_id": store_id})
        if not df.empty:
            records = [
                {
                    "store_id": store_id,
                    "official_sku_name": str(row["Official_SKU_Name"]),
                    "category": str(row.get("Category", "General")),
                    "default_unit": str(row.get("Default_Unit", "PCS")),
                    "gst_rate": float(row.get("GST_Rate", 18.0)),
                    "selling_price": float(row.get("Selling_Price", 0.0))
                }
                for _, row in df.iterrows()
            ]
            conn.execute(
                text("""
                    INSERT INTO master_skus (store_id, official_sku_name, category, default_unit, gst_rate, selling_price)
                    VALUES (:store_id, :official_sku_name, :category, :default_unit, :gst_rate, :selling_price)
                """),
                records
            )
    load_master.clear()

def add_single_sku_direct(store_slug: str, sku_name: str, category: str, unit: str, gst_rate: float, selling_price: float):
    engine = get_db_engine()
    store_id = get_or_create_store_id(store_slug)
    with engine.begin() as conn:
        conn.execute(
            text("""
                INSERT INTO master_skus (store_id, official_sku_name, category, default_unit, gst_rate, selling_price)
                VALUES (:store_id, :official_sku_name, :category, :default_unit, :gst_rate, :selling_price)
                ON CONFLICT (store_id, official_sku_name)
                DO UPDATE SET category = EXCLUDED.category, default_unit = EXCLUDED.default_unit, gst_rate = EXCLUDED.gst_rate, selling_price = EXCLUDED.selling_price
            """),
            {
                "store_id": store_id, "official_sku_name": sku_name, "category": category,
                "default_unit": unit, "gst_rate": gst_rate, "selling_price": selling_price
            }
        )
    load_master.clear()

def bulk_upsert_audited_skus(store_slug: str, records: list):
    if not records: return
    engine = get_db_engine()
    store_id = get_or_create_store_id(store_slug)
    with engine.begin() as conn:
        for rec in records:
            conn.execute(
                text("""
                    INSERT INTO master_skus (store_id, official_sku_name, category, default_unit, gst_rate, selling_price)
                    VALUES (:store_id, :official_sku_name, :category, :default_unit, :gst_rate, :selling_price)
                    ON CONFLICT (store_id, official_sku_name)
                    DO UPDATE SET category = EXCLUDED.category, default_unit = EXCLUDED.default_unit, gst_rate = EXCLUDED.gst_rate, selling_price = EXCLUDED.selling_price
                """),
                {
                    "store_id": store_id, "official_sku_name": rec["official"],
                    "category": rec["category"], "default_unit": rec["unit"],
                    "gst_rate": rec["gst"], "selling_price": rec["sp"]
                }
            )
    load_master.clear()

def delete_multiple_skus(store_slug: str, sku_list: list):
    if not sku_list: return
    engine = get_db_engine()
    store_id = get_or_create_store_id(store_slug)
    with engine.begin() as conn:
        conn.execute(
            text("DELETE FROM master_skus WHERE store_id = :store_id AND official_sku_name IN :sku_names"),
            {"store_id": store_id, "sku_names": tuple(sku_list)}
        )
    load_master.clear()

master_df = load_master(selected_store_slug)
master_sku_list = master_df["Official_SKU_Name"].dropna().tolist() if not master_df.empty else []
mapping_memory = load_json_memory(selected_store_slug)

# --- SKU MATCHING ENGINE ---
def match_sku(raw_name):
    cleaned_raw = raw_name.strip().upper()
    if cleaned_raw in mapping_memory: return mapping_memory[cleaned_raw]
    if master_sku_list:
        match, score, _ = process.extractOne(raw_name, master_sku_list, processor=utils.default_process)
        if score > 85: return match
        elif score >= 60: return f"⚠️ Needs Review: {match}"
    return raw_name

def get_known_selling_price(sku_name):
    clean_sku = sku_name.replace("⚠️ Needs Review: ", "").strip()
    if not master_df.empty and "Selling_Price" in master_df.columns:
        matched = master_df[master_df["Official_SKU_Name"] == clean_sku]
        if not matched.empty:
            price = matched.iloc[0]["Selling_Price"]
            if pd.notnull(price) and float(price) > 0: return float(price)
    return 0.0

# --- FAIL-SAFE AI INVOICE OCR ENGINE ---
def is_server_error(exception):
    err_str = str(exception).lower()
    return "503" in err_str or "unavailable" in err_str or "overloaded" in err_str or "429" in err_str or "resourceexhausted" in err_str

@retry(stop=stop_after_attempt(2), wait=wait_exponential(multiplier=1, min=1, max=3), retry=retry_if_exception(is_server_error), reraise=True)
def _call_gemini_with_retry(client, model_name, contents, config):
    return client.models.generate_content(model=model_name, contents=contents, config=config)

def extract_invoice_data_multiformat(file_bytes, mime_type="image/jpeg"):
    file_hash = hashlib.md5(file_bytes).hexdigest()
    try:
        cache = getattr(st.session_state, "ocr_file_hash_cache", {})
        if file_hash in cache: return cache[file_hash]
    except Exception:
        pass

    try:
        if "pdf" in mime_type.lower() and PYMUPDF_AVAILABLE:
            doc = fitz.open(stream=file_bytes, filetype="pdf")
            page = doc[0]
            pix = page.get_pixmap(dpi=110)
            img = Image.open(BytesIO(pix.tobytes("jpeg")))
            doc.close()
        else:
            img = Image.open(BytesIO(file_bytes))

        if img.mode in ("RGBA", "P"): img = img.convert("RGB")
        img.thumbnail((800, 800), Image.Resampling.BILINEAR)
        buffer = BytesIO()
        img.save(buffer, format="JPEG", quality=75, optimize=True)
        buffer.seek(0)
        optimized_bytes = buffer.getvalue()
        del img
        gc.collect()

        contents = [Image.open(BytesIO(optimized_bytes))]
        prompt = """
        You are an enterprise financial OCR system for wholesale, retail, plywood, hardware, and building material invoices.

        CRITICAL EXTRACTION & GROUND-TRUTH RULES:
        1. "Supplier Company Name": Main vendor/seller title from bill top header.
        2. "Invoice Number": Invoice/Bill number string if present, else "".
        3. "Invoice Date": Date string if present, else "".
        4. "Line Items": Extract every product row accurately.
           - "Item Name": Full product title or description. Read handwritten notes and pen edits carefully.
           - "Primary Quantity": Pure numeric count of physical pieces/sheets/boxes received (e.g. 6.0, 1.0, 70.0).
           - "Unit": Unit string (PCS, SET, SQM, SQFT, BOX, KG, LTR, NOS). Default "PCS".
           - "Printed Taxable Amount": Net Taxable Base value before tax after applying trade discounts.
           - "GST Rate": Total GST percentage as a pure number (0, 5, 12, 18, 28). Default 18.0.
           - "HSN Code": HSN/SAC code as string. If missing, "".

        OUTPUT SCHEMA (STRICT JSON ONLY):
        {
            "Supplier Company Name": "Vendor Title",
            "Invoice Number": "",
            "Invoice Date": "",
            "Line Items": [
                {
                    "Item Name": "Description",
                    "Primary Quantity": 1.0,
                    "Unit": "PCS",
                    "Printed Taxable Amount": 0.0,
                    "GST Rate": 18.0,
                    "HSN Code": ""
                }
            ]
        }
        """
        contents.append(prompt)
        config = types.GenerateContentConfig(response_mime_type="application/json")
        candidate_models = ['gemini-2.5-flash-lite', 'gemini-3.5-flash-lite', 'gemini-2.5-flash']
        shuffled_keys = list(API_KEYS_POOL)
        random.shuffle(shuffled_keys)

        for api_key in shuffled_keys:
            try:
                client = genai.Client(api_key=api_key)
                for model_name in candidate_models:
                    try:
                        response = _call_gemini_with_retry(client, model_name, contents, config)
                        text_res = response.text.strip()
                        if text_res.startswith("```json"): text_res = text_res[7:-3].strip()
                        elif text_res.startswith("```"): text_res = text_res[3:-3].strip()
                        parsed_res = json.loads(text_res)
                        try: st.session_state["ocr_file_hash_cache"][file_hash] = parsed_res
                        except Exception: pass
                        return parsed_res
                    except Exception: continue
            except Exception: continue
        raise Exception("All Gemini API keys failed or quota exhausted.")
    except Exception as e:
        return {"ERROR": str(e)}

def process_single_item_tuple(item_tuple):
    file_bytes, mime_type = item_tuple
    try: return extract_invoice_data_multiformat(file_bytes, mime_type)
    except Exception as e: return {"ERROR": str(e)}

# --- TAB 2 CATALOG PARSER HELPER (CHUNKED PDF OCR FOR 100+ PAGE SHEETS) ---
def parse_single_catalog_file(file_tuple):
    file_bytes, mime_type = file_tuple
    catalog_items = []
    try:
        if "pdf" in mime_type.lower() and PYMUPDF_AVAILABLE:
            pdf_doc = fitz.open(stream=file_bytes, filetype="pdf")
            total_pages = min(len(pdf_doc), 150)
            
            chunk_size = 5
            for start_idx in range(0, total_pages, chunk_size):
                end_idx = min(start_idx + chunk_size, total_pages)
                images_to_process = []
                
                for page_idx in range(start_idx, end_idx):
                    page = pdf_doc[page_idx]
                    pix = page.get_pixmap(dpi=120)
                    img = Image.open(BytesIO(pix.tobytes("jpeg")))
                    images_to_process.append(img)

                cat_prompt = """
                You are an enterprise catalog OCR system for building materials, laminates, door skins, plywood, and hardware.
                Extract the Supplier/Brand Company Name (e.g. DAARVI, GODREJ, GREENPLY) from the page header, and EVERY SINGLE product model and price row listed in the table.

                CRITICAL EXTRACTION RULES:
                1. "Supplier Name": Exact brand or company name publishing this price list. Standardize variations (e.g. "DA ARVI" -> "DAARVI").
                2. "Item Name": Full product title or description.
                3. "Model Code": Model number, item code, or series code if present, else "".
                4. "Unit": PCS, BOX, SET, LTR, KG, SQFT, MTR. Default "PCS".
                5. "GST Rate": GST percentage (0, 5, 12, 18, 28). Default 18.0.
                6. "Dealer Rate": Printed price/rate per piece as printed under RATE / PRICE column.

                OUTPUT SCHEMA (STRICT JSON ONLY):
                {
                    "Supplier Name": "Brand Name",
                    "Products": [
                        {
                            "Item Name": "",
                            "Model Code": "",
                            "Unit": "PCS",
                            "GST Rate": 18.0,
                            "Dealer Rate": 0.0
                        }
                    ]
                }
                """

                config = types.GenerateContentConfig(response_mime_type="application/json")
                candidate_models = ['gemini-2.5-flash-lite', 'gemini-3.5-flash-lite', 'gemini-2.5-flash']
                shuffled_keys = list(API_KEYS_POOL)
                random.shuffle(shuffled_keys)

                for img_obj in images_to_process:
                    if img_obj.mode in ("RGBA", "P"): img_obj = img_obj.convert("RGB")
                    img_obj.thumbnail((900, 900), Image.Resampling.BILINEAR)

                    buf = BytesIO()
                    img_obj.save(buf, format="JPEG", quality=80, optimize=True)
                    buf.seek(0)
                    raw_page_bytes = buf.getvalue()
                    
                    contents = [Image.open(BytesIO(raw_page_bytes)), cat_prompt]
                    parsed_page = None

                    for key in shuffled_keys:
                        try:
                            client = genai.Client(api_key=key)
                            for model_name in candidate_models:
                                try:
                                    res = _call_gemini_with_retry(client, model_name, contents, config)
                                    text_res = res.text.strip()
                                    if text_res.startswith("```json"): text_res = text_res[7:-3].strip()
                                    elif text_res.startswith("```"): text_res = text_res[3:-3].strip()
                                    parsed_page = json.loads(text_res)
                                    break
                                except Exception: continue
                            if parsed_page: break
                        except Exception: continue

                    if parsed_page:
                        raw_supplier = str(parsed_page.get("Supplier Name", "GENERIC VENDOR")).strip().upper()
                        supplier_name = "DAARVI" if "DA" in raw_supplier and "ARVI" in raw_supplier else raw_supplier
                        
                        for prod in parsed_page.get("Products", []):
                            item_name = str(prod.get("Item Name", "")).strip()
                            model_code = str(prod.get("Model Code", "")).strip()
                            full_title = f"{item_name} {model_code}".strip() if model_code and model_code not in item_name else item_name
                            printed_rate = float(prod.get("Dealer Rate") or 0.0)
                            if full_title and printed_rate >= 0:
                                catalog_items.append({
                                    "Supplier / Brand": supplier_name,
                                    "Model Name / Description": full_title,
                                    "Catalog Rate ₹": printed_rate,
                                    "GST Rate %": float(prod.get("GST Rate") or 18.0),
                                    "Unit": str(prod.get("Unit", "PCS")).upper()
                                })
                del images_to_process
                gc.collect()
            pdf_doc.close()
        else:
            img = Image.open(BytesIO(file_bytes))
            if img.mode in ("RGBA", "P"): img = img.convert("RGB")
            img.thumbnail((900, 900), Image.Resampling.BILINEAR)

            buf = BytesIO()
            img.save(buf, format="JPEG", quality=80, optimize=True)
            buf.seek(0)
            raw_page_bytes = buf.getvalue()
            
            cat_prompt = """Extract Supplier Name and product table rows. JSON ONLY: {"Supplier Name": "", "Products": [{"Item Name": "", "Model Code": "", "Unit": "PCS", "GST Rate": 18.0, "Dealer Rate": 0.0}]}"""
            contents = [Image.open(BytesIO(raw_page_bytes)), cat_prompt]
            config = types.GenerateContentConfig(response_mime_type="application/json")
            
            parsed_page = None
            for key in API_KEYS_POOL:
                try:
                    client = genai.Client(api_key=key)
                    res = _call_gemini_with_retry(client, 'gemini-2.5-flash', contents, config)
                    text_res = res.text.strip()
                    if text_res.startswith("```json"): text_res = text_res[7:-3].strip()
                    elif text_res.startswith("```"): text_res = text_res[3:-3].strip()
                    parsed_page = json.loads(text_res)
                    break
                except Exception: continue

            if parsed_page:
                raw_supplier = str(parsed_page.get("Supplier Name", "GENERIC VENDOR")).strip().upper()
                supplier_name = "DAARVI" if "DA" in raw_supplier and "ARVI" in raw_supplier else raw_supplier
                for prod in parsed_page.get("Products", []):
                    item_name = str(prod.get("Item Name", "")).strip()
                    model_code = str(prod.get("Model Code", "")).strip()
                    full_title = f"{item_name} {model_code}".strip() if model_code and model_code not in item_name else item_name
                    printed_rate = float(prod.get("Dealer Rate") or 0.0)
                    if full_title and printed_rate >= 0:
                        catalog_items.append({
                            "Supplier / Brand": supplier_name,
                            "Model Name / Description": full_title,
                            "Catalog Rate ₹": printed_rate,
                            "GST Rate %": float(prod.get("GST Rate") or 18.0),
                            "Unit": str(prod.get("Unit", "PCS")).upper()
                        })
    except Exception:
        pass
    return catalog_items

# --- REPORTLAB PDF GENERATOR ---
def generate_quotation_pdf(store_name: str, phone_str: str, customer_name: str, quote_df: pd.DataFrame, grand_total: float) -> bytes:
    if not REPORTLAB_AVAILABLE: raise ModuleNotFoundError("reportlab missing.")
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    story = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('DocTitle', parent=styles['Heading1'], fontSize=20, leading=24, textColor=colors.HexColor('#1E3A8A'), fontName='Helvetica-Bold', spaceAfter=4)
    subtitle_style = ParagraphStyle('DocSubtitle', parent=styles['Normal'], fontSize=9.5, leading=13, textColor=colors.HexColor('#4B5563'))
    meta_label = ParagraphStyle('MetaLabel', parent=styles['Normal'], fontSize=9, leading=12, textColor=colors.HexColor('#374151'))

    story.append(Paragraph("OFFICIAL QUOTATION", title_style))
    story.append(Paragraph(f"<b>Issued By:</b> {store_name.upper()}", subtitle_style))
    if phone_str: story.append(Paragraph(f"<b>Contact:</b> {phone_str}", subtitle_style))
    story.append(Spacer(1, 10))
    
    meta_data = [[Paragraph(f"<b>Customer Ref:</b> {customer_name}", meta_label), Paragraph(f"<b>Date:</b> {time.strftime('%d-%m-%Y %H:%M')}", meta_label)]]
    meta_table = Table(meta_data, colWidths=[310, 230])
    meta_table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F3F4F6')), ('PADDING', (0, 0), (-1, -1), 6)]))
    story.append(meta_table)
    story.append(Spacer(1, 14))

    table_data = [[Paragraph("S.No", styles['Normal']), Paragraph("Item Description", styles['Normal']), Paragraph("Qty", styles['Normal']), Paragraph("Unit", styles['Normal']), Paragraph("Final Rate (₹)", styles['Normal']), Paragraph("Total (₹)", styles['Normal'])]]
    for i, row in quote_df.iterrows():
        table_data.append([Paragraph(str(i + 1), styles['Normal']), Paragraph(str(row["Item Name"]), styles['Normal']), Paragraph(f"{float(row['Quantity']):g}", styles['Normal']), Paragraph(str(row["Unit"]), styles['Normal']), Paragraph(f"Rs. {float(row['Customer Unit Price (₹)']):,.2f}", styles['Normal']), Paragraph(f"Rs. {float(row['Total Value (₹)']):,.2f}", styles['Normal'])])

    item_table = Table(table_data, colWidths=[30, 240, 50, 50, 80, 90])
    item_table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')), ('TEXTCOLOR', (0, 0), (-1, 0), colors.white), ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB'))]))
    story.append(item_table)
    story.append(Spacer(1, 14))

    story.append(Paragraph(f"<b>Grand Total: Rs. {grand_total:,.2f}</b>", ParagraphStyle('GT', parent=styles['Heading2'], alignment=2, textColor=colors.HexColor('#1E3A8A'))))
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()

# --- WORKSPACE TABS ---
tab_parser, tab_catalog, tab_master, tab_memory, tab_guide = st.tabs([
    "📥 Batch Invoice Parser", 
    "📚 Price Catalog Extractor",
    "⚙️ Master Catalog",
    "📋 Vendor Memory", 
    "📖 Operating Guide"
])

# ==========================================
# TAB 1: BATCH INVOICE PARSER & QUOTATIONS
# ==========================================
with tab_parser:
    sm1, sm2, sm3 = st.columns(3)
    sm1.metric("Master SKUs Registered", len(master_sku_list))
    sm2.metric("Learned Vendor Rules", len(mapping_memory))
    sm3.metric("AI Engine Status", "🟢 Ready")

    st.divider()

    with st.expander("📄 Generate Customer Quotation (On-the-go)", expanded=False):
        st.caption("Create a professional PDF quote instantly either by manual quick entry or auto-calculating from an audited purchase bill.")
        saved_phone = st.session_state["user_store"].get("phone") or get_store_phone(selected_store_slug)
        q_tab_manual, q_tab_bill = st.tabs(["✍️ Quick Manual Entry", "📥 Auto-Quote from Parsed Bill"])
        
        with q_tab_manual:
            col_m1, col_m2 = st.columns([2, 2])
            with col_m1: man_cust_name = st.text_input("Customer Name", value="Walk-in Customer", key="man_cust_input")
            with col_m2: man_phone_input = st.text_input("Phone", value=saved_phone, key="man_phone_input")
                
            if man_phone_input.strip() != saved_phone.strip():
                save_store_phone(selected_store_slug, man_phone_input.strip())
                st.session_state["user_store"]["phone"] = man_phone_input.strip()

            col_mk1, col_mk2 = st.columns([1, 1])
            with col_mk1: manual_markup_pct = st.number_input("Profit Markup %", min_value=0.0, value=0.0, step=1.0, key="man_markup_input")
            with col_mk2: manual_disc_pct = st.number_input("Discount %", min_value=0.0, max_value=100.0, value=0.0, step=1.0, key="man_disc_input")

            if "manual_quote_data" not in st.session_state:
                st.session_state["manual_quote_data"] = pd.DataFrame([{"Item Name": "Plywood 18mm Commercial", "Quantity": 2.0, "Unit": "PCS", "MRP / Base Rate (₹)": 1800.0}])
                
            edited_manual_df = st.data_editor(st.session_state["manual_quote_data"], num_rows="dynamic", use_container_width=True, key="manual_quote_editor")
            st.session_state["manual_quote_data"] = edited_manual_df.copy()
            
            calc_manual_df = edited_manual_df.copy()
            calc_manual_df["Quantity"] = pd.to_numeric(calc_manual_df["Quantity"], errors='coerce').fillna(1.0)
            calc_manual_df["MRP (₹)"] = pd.to_numeric(calc_manual_df.get("MRP / Base Rate (₹)", 0.0), errors='coerce').fillna(0.0)
            calc_manual_df["Discount (%)"] = manual_disc_pct
            calc_manual_df["Customer Unit Price (₹)"] = (calc_manual_df["MRP (₹)"] * (1 + (manual_markup_pct / 100)) * (1 - (manual_disc_pct / 100))).round(2)
            calc_manual_df["Total Value (₹)"] = (calc_manual_df["Quantity"] * calc_manual_df["Customer Unit Price (₹)"]).round(2)
            
            man_grand_total = calc_manual_df["Total Value (₹)"].sum()
            st.metric("Quotation Total", f"₹{man_grand_total:,.2f}")
            
            if st.button("📄 Generate PDF Quotation", type="primary", use_container_width=True, key="btn_man_quote"):
                if not calc_manual_df.empty and man_grand_total > 0 and REPORTLAB_AVAILABLE:
                    pdf_bytes = generate_quotation_pdf(st.session_state['user_store']['display_name'], man_phone_input.strip(), man_cust_name, calc_manual_df, man_grand_total)
                    st.download_button("⬇️ Download PDF Quote", data=pdf_bytes, file_name=f"Quotation_{man_cust_name.replace(' ', '_')}.pdf", mime="application/pdf", type="primary", use_container_width=True)

    st.divider()

    col_upload, col_info = st.columns([2, 1])
    with col_upload:
        with st.container(border=True):
            st.markdown('<h3><span class="uos-num-badge">1</span>Ingestion Dropzone</h3>', unsafe_allow_html=True)
            c_mode, c_tax = st.columns([1, 1])
            with c_mode: upload_mode = st.radio("Input Mode:", ["📸 Direct Camera Snap", "📁 File Upload"], horizontal=True, key="dropzone_mode_switch")
            with c_tax: gst_bill_type = st.radio("Bill Tax Status:", ["Taxable GST Bill", "Non-GST / Net Rate Bill (0% Tax)"], horizontal=True, key="gst_bill_type_toggle")
            
            staged_file_tuples = []
            if upload_mode == "📸 Direct Camera Snap":
                cam_photo = st.camera_input("Take a photo of the purchase bill", key="direct_cam_input")
                if cam_photo: staged_file_tuples.append((cam_photo.getvalue(), "image/jpeg"))
            else:
                uploaded_files = st.file_uploader("Upload Bills", type=["jpg", "jpeg", "png", "pdf"], accept_multiple_files=True, label_visibility="collapsed", key="batch_file_uploader")
                if uploaded_files:
                    for f in uploaded_files:
                        staged_file_tuples.append((f.read(), "application/pdf" if f.name.lower().endswith(".pdf") else "image/jpeg"))

    with col_info:
        with st.container(border=True):
            st.subheader("⚡ Ingestion Queue")
            if staged_file_tuples: st.success(f"📁 **{len(staged_file_tuples)} File(s)** Ready")
            else: st.info("No files staged.")

    if staged_file_tuples:
        st.write("")
        c_btn1, c_btn2 = st.columns([3, 1])
        with c_btn1: process_btn = st.button("🚀 Process Invoices with AI Engine", type="primary", use_container_width=True)
        with c_btn2:
            if st.button("🧹 Clear Queue", use_container_width=True):
                if "parsed_df" in st.session_state: del st.session_state["parsed_df"]
                st.rerun()

        if process_btn:
            if "parsed_df" in st.session_state: del st.session_state["parsed_df"]
            all_parsed_items = []
            
            with st.status("Parsing purchase bills concurrently with AI...", expanded=True) as status_container:
                with ThreadPoolExecutor(max_workers=min(len(staged_file_tuples), 10)) as executor:
                    raw_results = list(executor.map(process_single_item_tuple, staged_file_tuples))
                    
                for parsed_json in raw_results:
                    if "ERROR" in parsed_json: continue
                    supplier = parsed_json.get("Supplier Company Name", "Unknown Supplier")
                    inv_num = parsed_json.get("Invoice Number", "")
                    inv_date = parsed_json.get("Invoice Date", "")
                    
                    if inv_num and is_duplicate_invoice_db(ACTIVE_STORE_ID, supplier, inv_num):
                        st.warning(f"⚠️ LEGAL AUDIT NOTICE: Invoice '{inv_num}' from '{supplier}' was previously registered in your store database.")

                    for row in parsed_json.get("Line Items", []):
                        qty = float(row.get("Primary Quantity") or 1.0)
                        if qty <= 0: qty = 1.0
                        gst_rate = 0.0 if gst_bill_type == "Non-GST / Net Rate Bill (0% Tax)" else float(row.get("GST Rate") or 18.0)
                        printed_taxable = float(row.get("Printed Taxable Amount") or 0.0)
                        base_rate_per_pc = printed_taxable / qty if qty > 0 else 0.0
                        raw_item_name = str(row.get("Item Name", "")).strip()
                        matched_sku = match_sku(raw_item_name)
                        
                        all_parsed_items.append({
                            "Supplier Name": supplier, "Invoice No": inv_num, "Invoice Date": inv_date,
                            "Raw Vendor Item": raw_item_name, "Official SKU": matched_sku, "Current Quantity": qty,
                            "Unit": str(row.get("Unit", "PCS")).upper(), "HSN/SAC": str(row.get("HSN Code") or "").strip(),
                            "Category": "General", "GST Rate": gst_rate, "Purchase Price": round(base_rate_per_pc, 2),
                            "Selling Price": round(get_known_selling_price(matched_sku), 2)
                        })
                status_container.update(label="✅ Ingestion Complete!", state="complete", expanded=False)
                
            if all_parsed_items:
                st.session_state["parsed_df"] = pd.DataFrame(all_parsed_items)
                gc.collect()
                st.rerun()

    if "parsed_df" in st.session_state:
        st.divider()
        st.markdown('<h3><span class="uos-num-badge">2</span>Live Inventory Audit Workspace</h3>', unsafe_allow_html=True)
        df = st.session_state["parsed_df"]
        
        # DETERMINISTIC ACCOUNTING MATH ENGINE
        df["Current Quantity"] = pd.to_numeric(df["Current Quantity"], errors='coerce').fillna(1.0)
        df["Purchase Price"] = pd.to_numeric(df["Purchase Price"], errors='coerce').fillna(0.0)
        df["GST Rate"] = pd.to_numeric(df["GST Rate"], errors='coerce').fillna(18.0)
        df["Selling Price"] = pd.to_numeric(df["Selling Price"], errors='coerce').fillna(0.0)
        
        df["Line Total (Excl. GST)"] = (df["Purchase Price"] * df["Current Quantity"]).round(2)
        df["GST Tax Amount"] = (df["Line Total (Excl. GST)"] * (df["GST Rate"] / 100.0)).round(2)
        df["Line Total (Incl. GST)"] = (df["Line Total (Excl. GST)"] + df["GST Tax Amount"]).round(2)
        df["Unit Cost (GST Paid) ₹"] = (df["Line Total (Incl. GST)"] / df["Current Quantity"]).round(2)
        
        total_taxable = df["Line Total (Excl. GST)"].sum()
        total_gst = df["GST Tax Amount"].sum()
        subtotal_incl_tax = total_taxable + total_gst

        uom_groups = df.groupby("Unit")["Current Quantity"].sum()
        uom_summary_str = " | ".join([f"{val:,.2f} {unit}" for unit, val in uom_groups.items()])

        c_m1, c_m2, c_m3, c_m4 = st.columns([1, 1, 1, 1])
        c_m1.metric("Total Line Items", f"{len(df)} Items")
        c_m2.metric("Stock Quantities by UOM", uom_summary_str)
        c_m3.metric("Taxable Base (Excl. GST)", f"₹{total_taxable:,.2f}")
        
        round_off_val = st.sidebar.number_input("Bill Round-Off Adjustment (₹)", value=0.0, step=0.05, format="%.2f")
        final_bill_total = subtotal_incl_tax + round_off_val
        c_m4.metric("Grand Total (GST Paid)", f"₹{final_bill_total:,.2f}", delta=f"GST Tax: ₹{total_gst:,.2f}")
        
        st.write("")
        display_columns = ["Raw Vendor Item", "Official SKU", "Current Quantity", "Unit", "HSN/SAC", "Purchase Price", "Line Total (Excl. GST)", "GST Rate", "Unit Cost (GST Paid) ₹", "Line Total (Incl. GST)", "Selling Price"]
        
        edited_display_df = st.data_editor(
            df[display_columns],
            num_rows="dynamic",
            use_container_width=True,
            key="audit_editor",
            column_config={
                "Raw Vendor Item": st.column_config.TextColumn("Raw Vendor Item", disabled=True),
                "Official SKU": st.column_config.SelectboxColumn("Official SKU Name", options=master_sku_list, required=True) if master_sku_list else "Official SKU",
                "Current Quantity": st.column_config.NumberColumn("Qty", min_value=0.01, format="%.2f"),
                "Purchase Price": st.column_config.NumberColumn("Unit Rate (Excl. GST) ₹", format="₹%.2f"),
                "Line Total (Excl. GST)": st.column_config.NumberColumn("Total (Excl. GST) ₹", format="₹%.2f", disabled=True),
                "GST Rate": st.column_config.NumberColumn("GST %", min_value=0, max_value=28, format="%d%%"),
                "Unit Cost (GST Paid) ₹": st.column_config.NumberColumn("Landed Cost (GST Paid) ₹/Pc", format="₹%.2f", disabled=True),
                "Line Total (Incl. GST)": st.column_config.NumberColumn("Total (Incl. GST) ₹", format="₹%.2f", disabled=True),
                "Selling Price": st.column_config.NumberColumn("Selling Price (SP) ₹", format="₹%.2f"),
            }
        )
        
        df_updated = edited_display_df.copy()
        df_updated["Current Quantity"] = pd.to_numeric(df_updated["Current Quantity"], errors='coerce').fillna(1.0)
        df_updated["Purchase Price"] = pd.to_numeric(df_updated["Purchase Price"], errors='coerce').fillna(0.0)
        df_updated["GST Rate"] = pd.to_numeric(df_updated["GST Rate"], errors='coerce').fillna(18.0)
        df_updated["Selling Price"] = pd.to_numeric(df_updated.get("Selling Price", 0.0), errors='coerce').fillna(0.0)

        df_updated["Line Total (Excl. GST)"] = (df_updated["Purchase Price"] * df_updated["Current Quantity"]).round(2)
        df_updated["GST Tax Amount"] = (df_updated["Line Total (Excl. GST)"] * (df_updated["GST Rate"] / 100.0)).round(2)
        df_updated["Line Total (Incl. GST)"] = (df_updated["Line Total (Excl. GST)"] + df_updated["GST Tax Amount"]).round(2)
        df_updated["Unit Cost (GST Paid) ₹"] = (df_updated["Line Total (Incl. GST)"] / df_updated["Current Quantity"]).round(2)
        
        st.session_state["parsed_df"] = df_updated

        st.divider()
        if st.button("✅ Confirm Audit & Generate Excel Import File", type="primary", use_container_width=True):
            memory_updated = False
            upsert_records = []
            
            for idx, row in df_updated.iterrows():
                raw = str(row.get("Raw Vendor Item", "")).strip().upper()
                official = str(row.get("Official SKU", "")).replace("⚠️ Needs Review: ", "").strip()
                cat_val = str(row.get("Category", "General"))
                unit_val = str(row.get("Unit", "PCS")).upper()
                gst_val = float(row.get("GST Rate", 18.0))
                sp_val = float(row.get("Selling Price", 0.0))
                
                v_supplier = str(row.get("Supplier Name", "UNKNOWN"))
                v_inv_no = str(row.get("Invoice No", ""))
                v_inv_date = str(row.get("Invoice Date", ""))
                log_invoice_to_db(ACTIVE_STORE_ID, v_supplier, v_inv_no, v_inv_date, final_bill_total)

                if raw and official and raw != official:
                    mapping_memory[raw] = official
                    memory_updated = True
                
                if official:
                    upsert_records.append({
                        "official": official, "category": cat_val,
                        "unit": unit_val, "gst": gst_val, "sp": sp_val
                    })

            if upsert_records: bulk_upsert_audited_skus(selected_store_slug, upsert_records)
            if memory_updated: save_json_memory(selected_store_slug, mapping_memory)
                
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Items"
            ws.cell(row=1, column=1, value=ACTIVE_STORE_DISPLAY)
            ws.cell(row=2, column=1, value="Items")
            ws.cell(row=3, column=1, value=f"Generated On: {time.strftime('%d-%m-%Y %H:%M:%S')}")
            
            exact_headers = ["S. No.", "Name", "Current Quantity", "Unit", "HSN/SAC", "Category", "GST Rate", "Selling Price", "Selling Price (Secondary)", "Purchase Price", "Purchase Price (Secondary)", "Secondary Unit", "Ratio"]
            for col_num, header_title in enumerate(exact_headers, 1): ws.cell(row=5, column=col_num, value=header_title)
            
            for i, row in df_updated.iterrows():
                row_idx = 6 + i
                raw_selling = row.get("Selling Price", 0.0)
                selling_val = float(raw_selling) if pd.notnull(raw_selling) and float(raw_selling) > 0 else ""
                purchase_val = float(row.get("Purchase Price", 0.0)) if pd.notnull(row.get("Purchase Price")) else 0.0
                clean_sku_name = str(row["Official SKU"]).replace("⚠️ Needs Review: ", "").strip()
                
                ws.cell(row=row_idx, column=1, value=i + 1)
                ws.cell(row=row_idx, column=2, value=clean_sku_name)
                ws.cell(row=row_idx, column=3, value=float(row["Current Quantity"]))
                ws.cell(row=row_idx, column=4, value=str(row["Unit"]).upper())
                ws.cell(row=row_idx, column=5, value=str(row.get("HSN/SAC", "")))
                ws.cell(row=row_idx, column=6, value=str(row.get("Category", "General")))
                ws.cell(row=row_idx, column=7, value=float(row["GST Rate"]))
                ws.cell(row=row_idx, column=8, value=selling_val)
                ws.cell(row=row_idx, column=9, value="")
                ws.cell(row=row_idx, column=10, value=purchase_val)
                ws.cell(row=row_idx, column=11, value="")
                ws.cell(row=row_idx, column=12, value="")
                ws.cell(row=row_idx, column=13, value="")

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            st.download_button(
                label=f"📥 Download ERP Bulk Import File for {ACTIVE_STORE_DISPLAY}",
                data=buffer.getvalue(),
                file_name=f"{selected_store_slug}_ERP_Stock_Import.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

# ==========================================
# TAB 2: PRICE CATALOG & MODEL EXTRACTOR
# ==========================================
with tab_catalog:
    st.subheader(f"📚 Multi-Vendor Price List & Catalog Batch Extractor ({ACTIVE_STORE_DISPLAY})")
    col_cat_left, col_cat_right = st.columns([2, 1])

    with col_cat_left:
        catalog_files = st.file_uploader("Upload Supplier Price Lists", type=["pdf", "png", "jpg", "jpeg"], accept_multiple_files=True, key="catalog_file_uploader_batch")
    with col_cat_right:
        catalog_tax_type = st.radio("Catalog Tax Status:", ["Taxable GST Rate", "Non-GST / Net Rate"], horizontal=True, key="cat_tax_status_radio")

    if catalog_files:
        st_files_tuples = [(f.read(), "application/pdf" if f.name.lower().endswith(".pdf") else "image/jpeg") for f in catalog_files]

        if st.button("🚀 Process All Price Lists Concurrently", type="primary", use_container_width=True, key="btn_run_catalog_ai"):
            all_extracted_catalog_items = []
            with st.status(f"Parsing {len(st_files_tuples)} price catalog(s) in parallel...", expanded=True) as status_box:
                try:
                    with ThreadPoolExecutor(max_workers=min(len(st_files_tuples), 6)) as executor:
                        results = list(executor.map(parse_single_catalog_file, st_files_tuples))
                        
                    for file_items in results:
                        for prod in file_items:
                            prod["GST Rate %"] = 0.0 if catalog_tax_type == "Non-GST / Net Rate" else float(prod.get("GST Rate %", 18.0))
                            all_extracted_catalog_items.append(prod)

                    status_box.update(label=f"✅ Extracted & Grouped {len(all_extracted_catalog_items)} total items!", state="complete", expanded=False)
                except Exception as cat_err:
                    status_box.update(label=f"❌ Failed to parse price lists: {cat_err}", state="error")

            if all_extracted_catalog_items:
                df_raw_cat = pd.DataFrame(all_extracted_catalog_items)
                df_sorted_cat = df_raw_cat.sort_values(by=["Supplier / Brand", "Model Name / Description"]).reset_index(drop=True)
                st.session_state["catalog_df"] = df_sorted_cat
                st.rerun()

    if "catalog_df" in st.session_state and st.session_state["catalog_df"] is not None and not st.session_state["catalog_df"].empty:
        st.divider()
        st.markdown("#### 📋 Extracted & Brand-Grouped Price Lists Preview")
        edited_cat_df = st.data_editor(st.session_state["catalog_df"], num_rows="dynamic", use_container_width=True, key="catalog_editor")

        c_cat1, c_cat2 = st.columns([1, 1])
        with c_cat1:
            if st.button("📥 Import Grouped Items into Master Catalog", type="primary", use_container_width=True, key="btn_import_catalog_master"):
                new_cat_records = []
                for _, row in edited_cat_df.iterrows():
                    sku_name = f"[{row.get('Supplier / Brand', 'GENERIC')}] {row['Model Name / Description']}".strip()
                    if sku_name:
                        new_cat_records.append({
                            "Official_SKU_Name": sku_name, "Category": str(row.get("Supplier / Brand", "General")),
                            "Default_Unit": str(row.get("Unit", "PCS")).upper(), "GST_Rate": float(row.get("GST Rate %", 18.0)),
                            "Selling_Price": float(row.get("Catalog Rate ₹", 0.0))
                        })
                if new_cat_records:
                    bulk_df = pd.DataFrame(new_cat_records)
                    combined = pd.concat([master_df, bulk_df], ignore_index=True).drop_duplicates(subset=["Official_SKU_Name"], keep="last")
                    save_master(combined, selected_store_slug)
                    st.toast(f"Imported {len(new_cat_records)} items to Master Catalog!")
                    st.rerun()

        with c_cat2:
            wb_cat = openpyxl.Workbook()
            ws_cat = wb_cat.active
            ws_cat.title = "Price Lists"
            ws_cat.cell(row=1, column=1, value=ACTIVE_STORE_DISPLAY)
            ws_cat.cell(row=2, column=1, value="Consolidated Supplier Price Lists")
            ws_cat.cell(row=3, column=1, value=f"Generated On: {time.strftime('%d-%m-%Y %H:%M:%S')}")

            exact_headers = ["S. No.", "Supplier / Brand", "Model Name / Description", "Catalog Rate ₹", "GST Rate %", "Unit"]
            for col_num, header_title in enumerate(exact_headers, 1): ws_cat.cell(row=5, column=col_num, value=header_title)

            for i, row in edited_cat_df.iterrows():
                row_idx = 6 + i
                ws_cat.cell(row=row_idx, column=1, value=i + 1)
                ws_cat.cell(row=row_idx, column=2, value=str(row.get("Supplier / Brand", "")).strip())
                ws_cat.cell(row=row_idx, column=3, value=str(row["Model Name / Description"]).strip())
                ws_cat.cell(row=row_idx, column=4, value=float(row.get("Catalog Rate ₹", 0.0)))
                ws_cat.cell(row=row_idx, column=5, value=float(row.get("GST Rate %", 18.0)))
                ws_cat.cell(row=row_idx, column=6, value=str(row.get("Unit", "PCS")).upper())

            buf_cat = BytesIO()
            wb_cat.save(buf_cat)
            buf_cat.seek(0)
            st.download_button("📥 Download Grouped Price List Excel File", data=buf_cat.getvalue(), file_name=f"{selected_store_slug}_Grouped_Price_Lists.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True, key="btn_dl_catalog_excel")

# ==========================================
# TAB 3: STORE MASTER CATALOG MANAGER
# ==========================================
with tab_master:
    st.subheader(f"⚙️ Master Inventory Catalog ({ACTIVE_STORE_DISPLAY})")
    
    with st.expander("📥 Bulk Import Catalog via Excel / CSV", expanded=False):
        uploaded_catalog = st.file_uploader("Upload Catalog (.xlsx or .csv)", type=["xlsx", "csv"], key="bulk_catalog_uploader")
        if uploaded_catalog is not None:
            try:
                imported_df = pd.read_csv(uploaded_catalog) if uploaded_catalog.name.endswith(".csv") else pd.read_excel(uploaded_catalog)
                st.dataframe(imported_df.head(5), use_container_width=True)
                col_sku = st.selectbox("Select SKU Name Column:", options=imported_df.columns)
                col_cat, col_unit, col_gst, col_price = st.columns(4)
                with col_cat: opt_cat = st.selectbox("Category Column:", options=["-- None --"] + list(imported_df.columns))
                with col_unit: opt_unit = st.selectbox("Unit Column:", options=["-- None --"] + list(imported_df.columns))
                with col_gst: opt_gst = st.selectbox("GST Column:", options=["-- None --"] + list(imported_df.columns))
                with col_price: opt_price = st.selectbox("Price Column:", options=["-- None --"] + list(imported_df.columns))
                
                if st.button("🚀 Import All SKUs into Master Catalog", type="primary", use_container_width=True):
                    new_records = []
                    for _, row in imported_df.iterrows():
                        sku_val = str(row[col_sku]).strip() if pd.notnull(row[col_sku]) else ""
                        if not sku_val or sku_val.lower() == "nan": continue
                        cat_val = str(row[opt_cat]).strip() if opt_cat != "-- None --" and pd.notnull(row[opt_cat]) else "General"
                        unit_val = str(row[opt_unit]).strip().upper() if opt_unit != "-- None --" and pd.notnull(row[opt_unit]) else "PCS"
                        try: gst_val = float(row[opt_gst]) if opt_gst != "-- None --" and pd.notnull(row[opt_gst]) else 18.0
                        except ValueError: gst_val = 18.0
                        try: price_val = float(row[opt_price]) if opt_price != "-- None --" and pd.notnull(row[opt_price]) else 0.0
                        except ValueError: price_val = 0.0
                        
                        new_records.append({"Official_SKU_Name": sku_val, "Category": cat_val, "Default_Unit": unit_val, "GST_Rate": gst_val, "Selling_Price": price_val})
                    
                    if new_records:
                        bulk_df = pd.DataFrame(new_records)
                        combined = pd.concat([master_df, bulk_df], ignore_index=True).drop_duplicates(subset=["Official_SKU_Name"], keep="last")
                        save_master(combined, selected_store_slug)
                        st.success("Successfully imported SKUs!")
                        st.rerun()
            except Exception as err:
                st.error(f"Error parsing file: {err}")

    st.divider()

    col_add, col_list = st.columns([1, 2])
    with col_add:
        with st.container(border=True):
            st.markdown("#### ➕ Add Single Master SKU")
            add_sku = st.text_input("SKU Name")
            add_cat = st.text_input("Category", value="General")
            add_unit = st.selectbox("Default Unit", options=["PCS", "BOX", "LTR", "KG", "NOS", "SET"])
            add_gst = st.selectbox("GST Rate (%)", options=[0, 5, 12, 18, 28], index=3)
            add_price = st.number_input("Selling Price ₹", min_value=0.0, step=10.0)
            
            if st.button("⚡ Save SKU to Catalog", use_container_width=True, type="primary"):
                if add_sku.strip():
                    add_single_sku_direct(selected_store_slug, add_sku.strip(), add_cat, add_unit, float(add_gst), float(add_price))
                    st.toast("Saved SKU!")
                    st.rerun()
                    
    with col_list:
        with st.container(border=True):
            st.markdown("#### 📋 Catalog Register")
            if not master_df.empty:
                search_query = st.text_input("🔍 Search Catalog SKUs...", key="catalog_search_bar")
                df_catalog_view = master_df.copy()
                if search_query.strip():
                    q = search_query.strip().lower()
                    mask = (df_catalog_view["Official_SKU_Name"].astype(str).str.lower().str.contains(q) | df_catalog_view["Category"].astype(str).str.lower().str.contains(q))
                    df_catalog_view = df_catalog_view[mask]

                df_catalog_view.insert(0, "Del", False)
                edited_catalog_df = st.data_editor(df_catalog_view, num_rows="fixed", hide_index=True, use_container_width=True, key="master_catalog_editor")
                selected_for_delete = edited_catalog_df[edited_catalog_df["Del"] == True]["Official_SKU_Name"].tolist()
                
                if selected_for_delete and st.button(f"🗑️ Delete Selected ({len(selected_for_delete)} SKUs)", type="primary", use_container_width=True):
                    delete_multiple_skus(selected_store_slug, selected_for_delete)
                    st.rerun()
            else:
                st.info("Master catalog for this store is currently empty.")

# ==========================================
# TAB 4: VENDOR SKU MEMORY WORKSPACE
# ==========================================
with tab_memory:
    st.subheader(f"🧠 Learned AI Vendor Memory ({ACTIVE_STORE_DISPLAY})")
    if mapping_memory:
        mem_df = pd.DataFrame([{"Raw Vendor Item Description": k, "Mapped Store SKU": v} for k, v in mapping_memory.items()])
        st.dataframe(mem_df, use_container_width=True)
        st.write("")
        if st.button("🗑️ Reset Store Memory Cache"):
            engine = get_db_engine()
            store_id = get_or_create_store_id(selected_store_slug)
            with engine.begin() as conn:
                conn.execute(text("DELETE FROM vendor_mappings WHERE store_id = :store_id"), {"store_id": store_id})
            load_json_memory.clear()
            st.success("Memory cache reset!")
            st.rerun()
    else:
        st.info("No learned vendor mappings recorded yet.")

# ==========================================
# TAB 5: IMPORT GUIDE
# ==========================================
with tab_guide:
    st.subheader("📖 Standard Operating Procedure")
    with st.container(border=True):
        st.markdown("""
        ### How to Process & Sync Invoices:
        1. **Select Store:** Choose active store in the sidebar directory.
        2. **Upload Bills:** Drop images, PDFs, or snap a photo directly in **Tab 1**.
        3. **Run AI Engine:** Click **Process Invoices** to extract line items.
        4. **Audit Workspace:** Check quantities, rates, and mapped SKUs.
        5. **Download Import File:** Generate the `.xlsx` spreadsheet for your ERP.
        """)
